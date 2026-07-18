"""Ingest monthly index-level series from MSCI exports (local xlsx) or the MSCI data service.

Indexes are driven by the registry (data/index_registry.csv) — this module iterates its rows
rather than scanning folders blindly, so the set of tracked indexes is explicit and editable.
Two sources:

- ``msci_local`` / ``msci_local_webweights``: the hand-downloaded '...Monthly.xlsx' exports
  (header row, then month-end levels).
- ``msci_api``: MSCI's own end-of-day data service (app2.msci.com, the same source behind the
  xlsx exports — verified identical to 9 significant figures on USA Momentum). ``returns_file``
  holds the numeric index code. Responses are cached to ``data/raw/msci_api/<code>.json``
  (committed, like the xlsx files) so offline runs and reproducibility keep working; on a
  network failure the cache is used, and only cache-AND-fetch failure raises.

We emit two tidy files:

  returns_monthly_long.csv : date, index_id, index_name, region, factor_type, level, ret
  levels_wide.csv          : date index x one column per series ("<region> | <factor_type>")

Run:  python -m portfolio_lab.ingest.returns
"""
from __future__ import annotations
import json
import re
import ssl
import urllib.request

import certifi
import openpyxl
import pandas as pd

from portfolio_lab import config as C

_DATE_RE = re.compile(r"\d{4}-\d{2}-\d{2}")

_MSCI_API = ("https://app2.msci.com/products/service/index/indexmaster/getLevelDataForGraph?"
             "currency_symbol=USD&index_variant=NETR&start_date=19690101&end_date={end}"
             "&data_frequency=END_OF_MONTH&index_codes={code}")


def _read_levels_api(code: str) -> list:
    """[(date, level)] from the MSCI data service, with a committed JSON cache."""
    cache = C.MSCI_API_CACHE_DIR / f"{code}.json"
    payload = None
    try:
        ctx = ssl.create_default_context(cafile=certifi.where())
        end = pd.Timestamp.today().strftime("%Y%m%d")
        start = "19690101"
        for _ in range(2):                                 # the service rejects dates before the
            url = _MSCI_API.format(code=code, end=end).replace("19690101", start)
            with urllib.request.urlopen(url, context=ctx, timeout=60) as r:
                payload = json.loads(r.read().decode())    # index's own inception and SAYS SO —
            if payload.get("indexes", {}).get("INDEX_LEVELS"):
                break                                      # retry once with its suggested date
            m = re.search(r"cannot be earlier than (\d{8})", str(payload.get("error_message", "")))
            payload = None
            if not m:
                break
            start = m.group(1)
        if payload:
            cache.parent.mkdir(parents=True, exist_ok=True)
            cache.write_text(json.dumps(payload))
    except Exception as e:
        print(f"[returns] WARN msci_api fetch failed for {code} ({e}) — trying cache")
        payload = None
    if payload is None:
        if not cache.exists():
            raise RuntimeError(f"msci_api index {code}: no network and no cache at {cache}")
        payload = json.loads(cache.read_text())
    out = []
    for row in payload["indexes"]["INDEX_LEVELS"]:
        d = str(row["calc_date"])
        out.append((f"{d[:4]}-{d[4:6]}-{d[6:8]}", float(row["level_eod"])))
    return out


def _read_levels(path):
    """Return [(date, level)] from an MSCI monthly workbook (skips header + non-date rows)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    data = list(ws.iter_rows(values_only=True))
    hdr = next(i for i, r in enumerate(data) if r and r[0] == "Date")
    out = []
    for r in data[hdr + 1:]:
        if not r or r[0] is None or r[1] is None:
            continue
        d = str(r[0])[:10]
        if _DATE_RE.match(d):
            out.append((d, float(r[1])))
    return out


def run() -> pd.DataFrame:
    C.ensure_dirs()
    rows = []
    for idx in C.load_registry():
        if idx["source"] == "msci_api":
            levels = _read_levels_api(idx["returns_file"])
        else:
            levels = _read_levels(C.RAW_DIR / idx["region"] / idx["returns_file"])
        for d, lvl in levels:
            rows.append([d, idx["index_id"], idx["display_name"],
                         idx["region"], idx["factor_type"], lvl])

    df = pd.DataFrame(rows, columns=["date", "index_id", "index_name",
                                     "region", "factor_type", "level"])
    df = df.sort_values(["region", "factor_type", "date"]).reset_index(drop=True)
    df["ret"] = df.groupby(["region", "factor_type"])["level"].pct_change()
    df.to_csv(C.RETURNS_LONG, index=False)

    df["col"] = df["region"] + " | " + df["factor_type"]
    wide = df.pivot_table(index="date", columns="col", values="level")
    wide.to_csv(C.LEVELS_WIDE)

    print(f"[returns] {len(df)} rows, {wide.shape[1]} series, "
          f"{wide.index.min()}..{wide.index.max()} -> {C.RETURNS_LONG.name}, {C.LEVELS_WIDE.name}")
    return df


if __name__ == "__main__":
    run()
